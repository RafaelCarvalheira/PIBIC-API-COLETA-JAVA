"""
Gerador de Relatorio Excel para resultados CE (Coleta e Entrega).

Gera 4 abas:
1. run_CE           - Com C8, com colaboracao
2. run_CE-NoShare   - Com C8, sem colaboracao
3. run_CEc8         - Sem C8, com colaboracao
4. run_CEc8-NoShare - Sem C8, sem colaboracao

Uso:
    python gerar_relatorio_ce.py [pasta_resultados] [pasta_json_input] [saida.xlsx]
"""

import json
import os
import re
import sys
from typing import Dict, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao encontrado. Instale com: pip install openpyxl")
    sys.exit(1)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "resultados_ce")
JSON_INPUT_DIR = os.path.join(SCRIPT_DIR, "json_output_ce")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "relatorio_ce.xlsx")


def load_json_file(filepath: str) -> Optional[Dict]:
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, dict) else None
    except Exception as e:
        print(f"  Erro ao ler {filepath}: {e}")
        return None


def extract_id_from_filename(filename: str) -> str:
    m = re.match(r'^(?:ce_)?(\d+)_', filename)
    return m.group(1) if m else ''


def get_num_customers(problem_id: str, input_dir: str) -> int:
    patterns = [
        f"vrps_{problem_id}_ce.json",
        f"{problem_id}_ce.json",
        f"vrps_{problem_id}L_ce.json",
        f"{problem_id}L_ce.json",
    ]
    for pattern in patterns:
        filepath = os.path.join(input_dir, pattern)
        if os.path.exists(filepath):
            data = load_json_file(filepath)
            if data:
                if 'globalParameters' in data:
                    n = data['globalParameters'].get('numberOfCustomers', 0)
                    if n: return n
                if 'customers' in data:
                    return len(data['customers'])
    return 0


def carrier_costs(result: Dict) -> Tuple[float, float]:
    a, b = 0.0, 0.0
    for route in result.get('routes', []):
        vid = route.get('vehicleId', '')
        c = route.get('routeCost', 0.0)
        if 'vehicle_1' in vid:
            a += c
        elif 'vehicle_2' in vid:
            b += c
    return a, b


def find_result_files(results_dir: str) -> Dict[str, Dict[str, str]]:
    """Organiza arquivos por ID. Chaves: 'ce', 'ce_no_share', 'cec8', 'cec8_no_share'."""
    result_files: Dict[str, Dict[str, str]] = {}
    if not os.path.isdir(results_dir):
        print(f"ERRO: pasta nao encontrada: {results_dir}")
        return {}

    for filename in sorted(os.listdir(results_dir)):
        if not filename.endswith('.json') or 'consolidado' in filename.lower():
            continue
        pid = extract_id_from_filename(filename)
        if not pid:
            continue
        filepath = os.path.join(results_dir, filename)
        bucket = result_files.setdefault(pid, {})

        # Ordem importa: checar sufixos mais especificos primeiro.
        if '_run_CEc8_no_share' in filename:
            bucket['cec8_no_share'] = filepath
        elif '_run_CE_no_share' in filename:
            bucket['ce_no_share'] = filepath
        elif '_run_CEc8' in filename:
            bucket['cec8'] = filepath
        elif '_run_CE' in filename:
            bucket['ce'] = filepath

    return result_files


# Estilos compartilhados
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
NUMBER_FMT = '#,##0.00'
FILL_CE = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")    # roxo para CE (com C8)
FILL_CEC8 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # azul para CEc8


def write_header(ws, headers, fill):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def write_simple_row(ws, row, pid, num_cust, cost, exec_time):
    ws.cell(row=row, column=1, value=int(pid) if pid.isdigit() else pid).border = THIN_BORDER
    ws.cell(row=row, column=2, value=num_cust).border = THIN_BORDER
    c = ws.cell(row=row, column=3, value=cost)
    c.number_format = NUMBER_FMT
    c.border = THIN_BORDER
    ws.cell(row=row, column=4, value=exec_time).border = THIN_BORDER


def write_noshare_row(ws, row, pid, num_cust, cost_a, cost_b, total, exec_time):
    ws.cell(row=row, column=1, value=int(pid) if pid.isdigit() else pid).border = THIN_BORDER
    ws.cell(row=row, column=2, value=num_cust).border = THIN_BORDER
    for col, val in ((3, cost_a), (4, cost_b), (5, total)):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = NUMBER_FMT
        c.border = THIN_BORDER
    ws.cell(row=row, column=6, value=exec_time).border = THIN_BORDER


def create_excel_report(results_dir: str, input_dir: str, output_path: str):
    print("=" * 60)
    print("GERADOR DE RELATORIO CE (4 abas)")
    print("=" * 60)

    result_files = find_result_files(results_dir)
    if not result_files:
        print("Nenhum resultado encontrado.")
        return
    print(f"Problemas encontrados: {len(result_files)}")

    wb = openpyxl.Workbook()

    ws_ce = wb.active
    ws_ce.title = "run_CE"
    write_header(ws_ce, ["ID", "Num Clientes", "OF (Custo)", "Tempo (s)"], FILL_CE)

    ws_ce_ns = wb.create_sheet("run_CE-NoShare")
    write_header(ws_ce_ns, ["ID", "Num Clientes", "OF Custo A", "OF Custo B", "OF Total", "Tempo (s)"], FILL_CE)

    ws_cec8 = wb.create_sheet("run_CEc8")
    write_header(ws_cec8, ["ID", "Num Clientes", "OF (Custo)", "Tempo (s)"], FILL_CEC8)

    ws_cec8_ns = wb.create_sheet("run_CEc8-NoShare")
    write_header(ws_cec8_ns, ["ID", "Num Clientes", "OF Custo A", "OF Custo B", "OF Total", "Tempo (s)"], FILL_CEC8)

    rows = {'ce': 2, 'ce_ns': 2, 'cec8': 2, 'cec8_ns': 2}

    def sort_key(x):
        try: return int(x)
        except ValueError: return float('inf')

    for pid in sorted(result_files.keys(), key=sort_key):
        files = result_files[pid]
        num_cust = get_num_customers(pid, input_dir)

        if 'ce' in files:
            d = load_json_file(files['ce'])
            if d:
                write_simple_row(ws_ce, rows['ce'], pid, num_cust,
                                 d.get('totalCost', 0.0), d.get('execution_time', ''))
                rows['ce'] += 1

        if 'ce_no_share' in files:
            d = load_json_file(files['ce_no_share'])
            if d:
                a, b = carrier_costs(d)
                write_noshare_row(ws_ce_ns, rows['ce_ns'], pid, num_cust,
                                  a, b, d.get('totalCost', 0.0), d.get('execution_time', ''))
                rows['ce_ns'] += 1

        if 'cec8' in files:
            d = load_json_file(files['cec8'])
            if d:
                write_simple_row(ws_cec8, rows['cec8'], pid, num_cust,
                                 d.get('totalCost', 0.0), d.get('execution_time', ''))
                rows['cec8'] += 1

        if 'cec8_no_share' in files:
            d = load_json_file(files['cec8_no_share'])
            if d:
                a, b = carrier_costs(d)
                write_noshare_row(ws_cec8_ns, rows['cec8_ns'], pid, num_cust,
                                  a, b, d.get('totalCost', 0.0), d.get('execution_time', ''))
                rows['cec8_ns'] += 1

    for ws in (ws_ce, ws_ce_ns, ws_cec8, ws_cec8_ns):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    print("=" * 60)
    print(f"Relatorio salvo em: {output_path}")
    print("=" * 60)


def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else RESULTS_DIR
    input_dir = sys.argv[2] if len(sys.argv) > 2 else JSON_INPUT_DIR
    output_path = sys.argv[3] if len(sys.argv) > 3 else OUTPUT_EXCEL
    create_excel_report(results_dir, input_dir, output_path)


if __name__ == "__main__":
    main()
