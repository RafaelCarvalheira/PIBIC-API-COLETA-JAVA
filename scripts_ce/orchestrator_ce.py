"""
Orquestrador CE - Roda instancias CE na API e gera relatorios.

Endpoints executados por instancia (nomes alinhados ao Julia supercodigo.jl):
- /run-CE: equivale ao run_CE do Julia (COM restricao C8, cada cliente 1 visita)
- /run-CEc8: equivale ao run_CEc8 do Julia (SEM C8, permite multiplas visitas)
- /run-CEc8-no-share: run_CEc8 sem colaboracao horizontal

Uso:
    python orchestrator_ce.py
    python orchestrator_ce.py --input <pasta_jsons> --output <pasta_saida>

Autor: PIBIC VRP Project
"""
import os
import sys
import json
import time
import csv
import requests
import argparse
from pathlib import Path
from datetime import datetime

# ============================================================
# CONFIGURACAO
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT_DIR = os.path.join(SCRIPT_DIR, "json_output_ce")
DEFAULT_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "resultados_ce")
# ============================================================

# Tentar importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("AVISO: openpyxl nao instalado. CSV sera gerado. Para Excel: pip install openpyxl")


# Configuracoes da API
API_BASE_URL = "http://localhost:8080/api/solve"
ENDPOINTS = {
    "run_CE":            f"{API_BASE_URL}/run-CE",             # COM C8 (= Julia run_CE)
    "run_CEc8":          f"{API_BASE_URL}/run-CEc8",           # SEM C8 (= Julia run_CEc8)
    "run_CEc8_no_share": f"{API_BASE_URL}/run-CEc8-no-share",
    "run_CE_no_share":   f"{API_BASE_URL}/run-CE-no-share",
}


def load_json_file(filepath: str) -> dict:
    """Carrega um arquivo JSON."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json_file(data: dict, filepath: str):
    """Salva dados em um arquivo JSON."""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def call_api(endpoint: str, input_data: dict) -> dict:
    """Faz uma chamada POST para a API."""
    try:
        response = requests.post(endpoint, json=input_data, timeout=300)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"    ERRO na API: {e}")
        return None


def process_instance(json_path: str, output_dir: str) -> dict:
    """
    Processa uma instancia CE: chama CE-CUSTOM, CE-CUSTOM-NO-SHARE e CE-C8.
    Retorna um dicionario com os resultados para o relatorio.
    """
    input_data = load_json_file(json_path)
    problem_id = input_data.get("problemId", "unknown")
    num_customers = input_data.get("globalParameters", {}).get("numberOfCustomers", 0)

    print(f"  Processando instancia {problem_id} ({num_customers} clientes)...")

    instance_start_time = time.time()

    results = {
        "problem_id": problem_id,
        "num_customers": num_customers,
        "input_file": os.path.basename(json_path),
    }

    # ========================================
    # 1. run_CEc8 (SEM C8 - permite multiplas visitas, com colaboracao)
    # ========================================
    print(f"    Chamando /run-CEc8...")
    start_time = time.time()
    cec8_result = call_api(ENDPOINTS["run_CEc8"], input_data)
    cec8_time = time.time() - start_time

    if cec8_result:
        results["run_CEc8"] = {
            "total_cost": cec8_result.get("totalCost", 0),
            "num_routes": len(cec8_result.get("routes", [])),
            "status": cec8_result.get("status", ""),
            "message": cec8_result.get("message", ""),
            "execution_time": round(cec8_time, 2),
        }
        cec8_result["execution_time"] = round(cec8_time, 2)
        cec8_output_path = os.path.join(output_dir, f"{problem_id}_run_CEc8.json")
        save_json_file(cec8_result, cec8_output_path)
        print(f"    run_CEc8: Custo = {cec8_result.get('totalCost', 0):.2f} ({cec8_time:.2f}s)")
    else:
        results["run_CEc8"] = {"error": "Falha na API"}

    # ========================================
    # 2. run_CEc8-no-share (SEM C8, sem colaboracao)
    # ========================================
    print(f"    Chamando /run-CEc8-no-share...")
    start_time = time.time()
    no_share_result = call_api(ENDPOINTS["run_CEc8_no_share"], input_data)
    no_share_time = time.time() - start_time

    if no_share_result:
        routes = no_share_result.get("routes", [])
        cost_a = sum(r.get("routeCost", 0) for r in routes if "vehicle_1" in r.get("vehicleId", ""))
        cost_b = sum(r.get("routeCost", 0) for r in routes if "vehicle_2" in r.get("vehicleId", ""))

        results["run_CEc8_no_share"] = {
            "total_cost": no_share_result.get("totalCost", 0),
            "cost_a": cost_a,
            "cost_b": cost_b,
            "num_routes": len(routes),
            "status": no_share_result.get("status", ""),
            "message": no_share_result.get("message", ""),
            "execution_time": round(no_share_time, 2),
        }
        no_share_result["execution_time"] = round(no_share_time, 2)
        no_share_output_path = os.path.join(output_dir, f"{problem_id}_run_CEc8_no_share.json")
        save_json_file(no_share_result, no_share_output_path)
        print(f"    run_CEc8-NO-SHARE: Custo = {no_share_result.get('totalCost', 0):.2f} (A={cost_a:.2f}, B={cost_b:.2f}) ({no_share_time:.2f}s)")
    else:
        results["run_CEc8_no_share"] = {"error": "Falha na API"}

    # ========================================
    # 3. run_CE (COM C8 - 1 visita por cliente, com colaboracao)
    # ========================================
    print(f"    Chamando /run-CE...")
    start_time = time.time()
    ce_result = call_api(ENDPOINTS["run_CE"], input_data)
    ce_time = time.time() - start_time

    if ce_result:
        results["run_CE"] = {
            "total_cost": ce_result.get("totalCost", 0),
            "num_routes": len(ce_result.get("routes", [])),
            "status": ce_result.get("status", ""),
            "message": ce_result.get("message", ""),
            "execution_time": round(ce_time, 2),
        }
        ce_result["execution_time"] = round(ce_time, 2)
        ce_output_path = os.path.join(output_dir, f"{problem_id}_run_CE.json")
        save_json_file(ce_result, ce_output_path)
        print(f"    run_CE: Custo = {ce_result.get('totalCost', 0):.2f} ({ce_time:.2f}s)")
    else:
        results["run_CE"] = {"error": "Falha na API"}

    # ========================================
    # 4. run_CE-no-share (COM C8, sem colaboracao)
    # ========================================
    print(f"    Chamando /run-CE-no-share...")
    start_time = time.time()
    ce_ns_result = call_api(ENDPOINTS["run_CE_no_share"], input_data)
    ce_ns_time = time.time() - start_time

    if ce_ns_result:
        routes = ce_ns_result.get("routes", [])
        cost_a = sum(r.get("routeCost", 0) for r in routes if "vehicle_1" in r.get("vehicleId", ""))
        cost_b = sum(r.get("routeCost", 0) for r in routes if "vehicle_2" in r.get("vehicleId", ""))

        results["run_CE_no_share"] = {
            "total_cost": ce_ns_result.get("totalCost", 0),
            "cost_a": cost_a,
            "cost_b": cost_b,
            "num_routes": len(routes),
            "status": ce_ns_result.get("status", ""),
            "message": ce_ns_result.get("message", ""),
            "execution_time": round(ce_ns_time, 2),
        }
        ce_ns_result["execution_time"] = round(ce_ns_time, 2)
        ce_ns_output_path = os.path.join(output_dir, f"{problem_id}_run_CE_no_share.json")
        save_json_file(ce_ns_result, ce_ns_output_path)
        print(f"    run_CE-NO-SHARE: Custo = {ce_ns_result.get('totalCost', 0):.2f} (A={cost_a:.2f}, B={cost_b:.2f}) ({ce_ns_time:.2f}s)")
    else:
        results["run_CE_no_share"] = {"error": "Falha na API"}

    # Tempo total da instancia
    instance_total_time = time.time() - instance_start_time
    results["instance_time"] = round(instance_total_time, 2)
    print(f"    TEMPO TOTAL DA INSTANCIA: {instance_total_time:.2f}s")

    return results


def create_excel_report(all_results: list, output_path: str):
    """Cria relatorio Excel com os resultados CE (3 endpoints)."""
    if not EXCEL_AVAILABLE:
        print("Excel nao disponivel. Salvando como JSON...")
        json_path = output_path.replace(".xlsx", "_report.json")
        save_json_file(all_results, json_path)
        return

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill_c8 = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================
    # ABA 1: CE-Custom (Com Colaboracao, Multi-Start)
    # ========================================
    ws_ce = wb.active
    ws_ce.title = "run_CEc8"

    ce_headers = ["ID", "Num. Clientes", "OF (Custo)", "Num. Rotas", "Tempo (s)"]
    for col, header in enumerate(ce_headers, 1):
        cell = ws_ce.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, result in enumerate(all_results, 2):
        ce_data = result.get("run_CEc8", {})
        ws_ce.cell(row=row, column=1, value=result.get("problem_id", "")).border = thin_border
        ws_ce.cell(row=row, column=2, value=result.get("num_customers", 0)).border = thin_border
        ws_ce.cell(row=row, column=3, value=round(ce_data.get("total_cost", 0), 2)).border = thin_border
        ws_ce.cell(row=row, column=4, value=ce_data.get("num_routes", 0)).border = thin_border
        ws_ce.cell(row=row, column=5, value=ce_data.get("execution_time", 0)).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_ce.column_dimensions[col].width = 14

    # ========================================
    # ABA 2: CE-Custom-NO-SHARE (Sem Colaboracao, Multi-Start)
    # ========================================
    ws_no_share = wb.create_sheet("run_CEc8-NoShare")

    no_share_headers = ["ID", "Num. Clientes", "OF-A", "OF-B", "OF Total", "Tempo (s)"]
    for col, header in enumerate(no_share_headers, 1):
        cell = ws_no_share.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, result in enumerate(all_results, 2):
        no_share_data = result.get("run_CEc8_no_share", {})
        ws_no_share.cell(row=row, column=1, value=result.get("problem_id", "")).border = thin_border
        ws_no_share.cell(row=row, column=2, value=result.get("num_customers", 0)).border = thin_border
        ws_no_share.cell(row=row, column=3, value=round(no_share_data.get("cost_a", 0), 2)).border = thin_border
        ws_no_share.cell(row=row, column=4, value=round(no_share_data.get("cost_b", 0), 2)).border = thin_border
        ws_no_share.cell(row=row, column=5, value=round(no_share_data.get("total_cost", 0), 2)).border = thin_border
        ws_no_share.cell(row=row, column=6, value=no_share_data.get("execution_time", 0)).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_no_share.column_dimensions[col].width = 14

    # ========================================
    # ABA 3: Comparacao CE-Custom
    # ========================================
    ws_compare = wb.create_sheet("Comparacao run_CEc8")

    compare_headers = ["ID", "Num. Clientes", "run_CEc8", "run_CEc8-NoShare", "Economia", "Economia (%)", "Tempo Instancia (s)"]
    for col, header in enumerate(compare_headers, 1):
        cell = ws_compare.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, result in enumerate(all_results, 2):
        ce_cost = result.get("run_CEc8", {}).get("total_cost", 0)
        no_share_cost = result.get("run_CEc8_no_share", {}).get("total_cost", 0)
        economia = no_share_cost - ce_cost
        economia_pct = (economia / no_share_cost * 100) if no_share_cost > 0 else 0

        ws_compare.cell(row=row, column=1, value=result.get("problem_id", "")).border = thin_border
        ws_compare.cell(row=row, column=2, value=result.get("num_customers", 0)).border = thin_border
        ws_compare.cell(row=row, column=3, value=round(ce_cost, 2)).border = thin_border
        ws_compare.cell(row=row, column=4, value=round(no_share_cost, 2)).border = thin_border
        ws_compare.cell(row=row, column=5, value=round(economia, 2)).border = thin_border
        ws_compare.cell(row=row, column=6, value=round(economia_pct, 2)).border = thin_border
        ws_compare.cell(row=row, column=7, value=result.get("instance_time", 0)).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_compare.column_dimensions[col].width = 15

    # ========================================
    # ABA 4: CE-C8 (Com Colaboracao, COM C8)
    # ========================================
    ws_ce_c8 = wb.create_sheet("run_CE (Com C8)")

    ce_c8_headers = ["ID", "Num. Clientes", "OF (Custo)", "Num. Rotas", "Tempo (s)"]
    for col, header in enumerate(ce_c8_headers, 1):
        cell = ws_ce_c8.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_c8
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, result in enumerate(all_results, 2):
        ce_c8_data = result.get("run_CE", {})
        ws_ce_c8.cell(row=row, column=1, value=result.get("problem_id", "")).border = thin_border
        ws_ce_c8.cell(row=row, column=2, value=result.get("num_customers", 0)).border = thin_border
        ws_ce_c8.cell(row=row, column=3, value=round(ce_c8_data.get("total_cost", 0), 2)).border = thin_border
        ws_ce_c8.cell(row=row, column=4, value=ce_c8_data.get("num_routes", 0)).border = thin_border
        ws_ce_c8.cell(row=row, column=5, value=ce_c8_data.get("execution_time", 0)).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_ce_c8.column_dimensions[col].width = 14

    # ========================================
    # ABA 5: run_CE-NoShare (COM C8, sem colaboracao)
    # ========================================
    ws_ce_ns = wb.create_sheet("run_CE-NoShare")

    ce_ns_headers = ["ID", "Num. Clientes", "OF-A", "OF-B", "OF Total", "Tempo (s)"]
    for col, header in enumerate(ce_ns_headers, 1):
        cell = ws_ce_ns.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_c8
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, result in enumerate(all_results, 2):
        ce_ns_data = result.get("run_CE_no_share", {})
        ws_ce_ns.cell(row=row, column=1, value=result.get("problem_id", "")).border = thin_border
        ws_ce_ns.cell(row=row, column=2, value=result.get("num_customers", 0)).border = thin_border
        ws_ce_ns.cell(row=row, column=3, value=round(ce_ns_data.get("cost_a", 0), 2)).border = thin_border
        ws_ce_ns.cell(row=row, column=4, value=round(ce_ns_data.get("cost_b", 0), 2)).border = thin_border
        ws_ce_ns.cell(row=row, column=5, value=round(ce_ns_data.get("total_cost", 0), 2)).border = thin_border
        ws_ce_ns.cell(row=row, column=6, value=ce_ns_data.get("execution_time", 0)).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_ce_ns.column_dimensions[col].width = 14

    # Salvar
    wb.save(output_path)
    print(f"\nRelatorio Excel CE salvo em: {output_path}")


def create_csv_report(all_results: list, output_path: str):
    """Cria relatorio CSV com os resultados CE (3 endpoints)."""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')

        # Cabecalho
        writer.writerow([
            'problem_id', 'num_customers',
            # run_CEc8
            'run_CEc8_cost', 'run_CEc8_routes', 'run_CEc8_time_s',
            # run_CEc8-NoShare
            'run_CEc8_no_share_cost', 'run_CEc8_no_share_cost_a', 'run_CEc8_no_share_cost_b', 'run_CEc8_no_share_time_s',
            # Economia run_CEc8
            'run_CEc8_economia', 'run_CEc8_economia_pct',
            # run_CE (com C8)
            'run_CE_cost', 'run_CE_routes', 'run_CE_time_s',
            # run_CE-NoShare (com C8, sem colaboracao)
            'run_CE_no_share_cost', 'run_CE_no_share_cost_a', 'run_CE_no_share_cost_b', 'run_CE_no_share_time_s',
            # Tempo total
            'instance_time_s'
        ])

        # Dados
        for result in all_results:
            ce_data = result.get('run_CEc8', {})
            no_share_data = result.get('run_CEc8_no_share', {})
            ce_c8_data = result.get('run_CE', {})
            ce_ns_data = result.get('run_CE_no_share', {})

            ce_cost = ce_data.get('total_cost', 0)
            no_share_cost = no_share_data.get('total_cost', 0)
            economia = no_share_cost - ce_cost
            economia_pct = (economia / no_share_cost * 100) if no_share_cost > 0 else 0

            writer.writerow([
                result.get('problem_id', ''),
                result.get('num_customers', 0),
                round(ce_cost, 2),
                ce_data.get('num_routes', 0),
                ce_data.get('execution_time', 0),
                round(no_share_cost, 2),
                round(no_share_data.get('cost_a', 0), 2),
                round(no_share_data.get('cost_b', 0), 2),
                no_share_data.get('execution_time', 0),
                round(economia, 2),
                round(economia_pct, 2),
                round(ce_c8_data.get('total_cost', 0), 2),
                ce_c8_data.get('num_routes', 0),
                ce_c8_data.get('execution_time', 0),
                round(ce_ns_data.get('total_cost', 0), 2),
                round(ce_ns_data.get('cost_a', 0), 2),
                round(ce_ns_data.get('cost_b', 0), 2),
                ce_ns_data.get('execution_time', 0),
                result.get('instance_time', 0)
            ])

    print(f"Relatorio CSV CE salvo em: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Orquestrador CE - Processa instancias CE")
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT_DIR,
                        help=f"Pasta contendo os arquivos JSON (default: {DEFAULT_INPUT_DIR})")
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT_DIR,
                        help=f"Pasta para salvar resultados (default: {DEFAULT_OUTPUT_DIR})")
    args = parser.parse_args()

    input_dir = args.input
    output_dir = args.output

    # Verificar pasta de entrada
    if not os.path.isdir(input_dir):
        print(f"ERRO: Pasta '{input_dir}' nao encontrada.")
        sys.exit(1)

    # Criar pasta de saida
    os.makedirs(output_dir, exist_ok=True)

    # Listar arquivos JSON
    json_files = sorted([f for f in os.listdir(input_dir) if f.endswith(".json")])

    if not json_files:
        print(f"ERRO: Nenhum arquivo JSON encontrado em '{input_dir}'")
        sys.exit(1)

    print(f"=" * 60)
    print(f"ORQUESTRADOR CE (run_CEc8, run_CEc8-no-share, run_CE)")
    print(f"=" * 60)
    print(f"Pasta de entrada: {input_dir}")
    print(f"Pasta de saida: {output_dir}")
    print(f"Instancias encontradas: {len(json_files)}")
    print(f"API: {API_BASE_URL}")
    print(f"=" * 60)

    # Verificar se a API esta rodando
    print("\nVerificando conexao com a API...")
    try:
        test_data = {"problemId": "test", "globalParameters": {"vehicleCapacity": 100}}
        requests.post(ENDPOINTS["run_CEc8"], json=test_data, timeout=5)
        print("API esta respondendo!\n")
    except requests.exceptions.ConnectionError:
        print("ERRO: Nao foi possivel conectar a API.")
        print("Certifique-se de que a API esta rodando: mvn spring-boot:run")
        sys.exit(1)
    except:
        print("API esta respondendo!\n")

    # Processar cada instancia
    all_results = []
    start_total = time.time()

    for i, json_file in enumerate(json_files, 1):
        print(f"\n[{i}/{len(json_files)}] {json_file}")
        json_path = os.path.join(input_dir, json_file)

        try:
            result = process_instance(json_path, output_dir)
            all_results.append(result)
        except Exception as e:
            print(f"  ERRO ao processar: {e}")
            all_results.append({
                "problem_id": json_file.replace(".json", ""),
                "error": str(e)
            })

    total_time = time.time() - start_total

    # Salvar resultados consolidados em JSON
    consolidated_path = os.path.join(output_dir, "resultados_consolidados_ce.json")
    save_json_file(all_results, consolidated_path)
    print(f"\nResultados consolidados salvos em: {consolidated_path}")

    # Gerar relatorio CSV
    csv_path = os.path.join(output_dir, "relatorio_vrp_ce.csv")
    create_csv_report(all_results, csv_path)

    # Gerar relatorio Excel
    if EXCEL_AVAILABLE:
        excel_path = os.path.join(output_dir, "relatorio_vrp_ce.xlsx")
        create_excel_report(all_results, excel_path)

    # Resumo final
    print(f"\n{'=' * 60}")
    print(f"RESUMO CE")
    print(f"{'=' * 60}")
    print(f"Instancias processadas: {len(all_results)}")
    print(f"Tempo total: {total_time:.2f} segundos ({total_time/60:.2f} minutos)")
    print(f"Tempo medio por instancia: {total_time/len(all_results):.2f} segundos")
    print(f"Resultados salvos em: {output_dir}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
